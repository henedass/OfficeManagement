"""'
written by
NIKA REZAEE
SADENEH SADEGHI
"""
import PySimpleGUI as sg
import pandas as pd
from openpyxl import Workbook
import openpyxl

# accessing the Excel database through openpyxl

wb = openpyxl.load_workbook('SWD_data.xlsx')
ws = wb.active

# OIDs list for maintaining all initial Personnel IDs in database
# PIDs list for maintaining all initial IDs(Personal IDs) in database

PIDs = []
OIDs = []

for i in range(ws.max_row - 1):
    PIDs.append(str(ws.cell(row=i + 2, column=3).value))
    OIDs.append(str(ws.cell(row=i + 2, column=4).value))

# list of all positions possible to have

list_of_positions = ['CEO', 'CTO', 'COO', 'CMO', 'Accountant']

# Welcome layout, created using PySimpleGUI, the first window that opens while running the app
# also the main menu called in main function
# has exit button, Personnel Management Panel for management access only, Personnel Panel for personnel access only

layout_welcome = [
    [sg.Text('Welcome To Main Menu', font=('Arial Bold', 40), expand_x=True, justification='center')],
    [sg.Button("Personnel Management Panel", size=(100, 5), pad=(400, 10))],
    [sg.Button("Personnel Panel", size=(100, 5), pad=(400, 10))],
    [sg.Button("Exit", size=(100, 5), pad=(400, 10))]
]


# login function, opening password window for personnel access to their personal page
# accessible for registered personnel
# protected by password and Personnel ID
# default password is personnel's personal ID (ID)

def login():
    layout_login = [
        [sg.Text('username', size=(15, 1)), sg.InputText()],
        [sg.Text('password', size=(15, 1)), sg.InputText()],
        [sg.Submit()]
    ]
    window = sg.Window("Log In", layout_login, size=(800, 300), modal=True)
    while True:
        event, values = window.read()  # Personnel ID and password save into values List
        if event == sg.WIN_CLOSED:
            break
        if event == "Submit":
            if values[0] in OIDs:  # checking if personnel ID exists and finding its index in OIDs saving as index
                index = OIDs.index(values[0])
                if values[1] == str(ws.cell(row=index + 2, column=3).value):  # checking if password is equal
                    # to that person's ID
                    personnel_panel(index)  # opening personnel panel of the person with entered personnel ID
                    break
                else:
                    sg.popup("Password Incorrect", "Password is wrong.")  # popup if password is wrong
            else:
                sg.popup("Invalid Input", "No one with this Personnel ID exists.")  # popup if personnel ID doesnt exist

    window.close()


# personnel panel function gets row as parameter, pointing to the specific personnel's row in excel database
# personnel panel window receives date and working hours and saves them into excel database

def personnel_panel(row):
    layout_personnel = [
        [sg.Text(
            f'Welcome to Personnel Panel {ws.cell(row=row + 2, column=1).value} {ws.cell(row=row + 2, column=2).value}',
            font=('Arial Bold', 20), expand_x=True, justification='center')],
        [sg.Text('Please enter your working hours and date below:')],
        [sg.Text('Day', size=(10, 1)), sg.InputText(size=(5, 5)), sg.Text('   /'), sg.Text('Month', size=(10, 1)),
         sg.InputText(size=(5, 5))],
        [sg.Text('Time Arriving:', size=(10, 1)), sg.InputText(size=(5, 5)), sg.Text('   :'), sg.Text('', size=(1, 1)),
         sg.InputText(size=(5, 5))],
        [sg.Text('Time Leaving:', size=(10, 1)), sg.InputText(size=(5, 5)), sg.Text('   :'), sg.Text('', size=(1, 1)),
         sg.InputText(size=(5, 5))],
        [sg.Submit()]
    ]
    window = sg.Window("Personnel Panel", layout_personnel, size=(800, 300), modal=True)
    while True:
        event, values = window.read()  # day, month, arrival hour, arrival minute, leaving hour and leaving minute gets
        # stored in values list
        if event == sg.WIN_CLOSED:
            break
        if event == "Submit":
            # constrains: day and month and hours and minutes should be all digit
            # date format is DD/MM and only available on Iranian calendar
            # working hours can be between 8:00 and 21:59
            # else popup will appear for each wrong input
            if values[0].isdigit() and values[1].isdigit() and 0 < int(values[0]) < 32 and 0 < int(values[1]) < 13:
                if (int(values[0]) == 31 and int(values[1]) > 6) or (int(values[0]) == 30 and int(values[1]) == 12):
                    sg.popup("Invalid Input", "Day or month may be wrong.")
                else:
                    if values[2].isdigit() and values[3].isdigit() and values[4].isdigit() and values[
                         5].isdigit() and 8 <= int(values[2]) < 22 and 8 <= int(values[4]) < 22 and 0 <= int(
                         values[3]) < 60 and 0 <= int(values[5]) < 60 and int(values[2]) <= int(values[4]):
                        worked_minutes = (int(values[4]) - int(values[2])) * 60 + int(values[5]) - int(
                            values[3])  # counting working minutes of the day
                        # if a working minutes for that day already exists, the second working minute will be replaced

                        # saving working minute for the personnel with date as column name
                        column_names = []  # saving column names into a list
                        for cell in ws[1]:
                            column_names.append(cell.value)

                        if f'{values[0]}/{values[1]}' in column_names:  # checking if the column of that date already
                            # exists: if YES: add the working minutes to correspondent cell - if NO: create a new column
                            # with the date as the name of the column and only add this personnel's working minutes
                            column = column_names.index(f'{values[0]}/{values[1]}')
                            ws.cell(row=row + 2, column=column + 1).value = worked_minutes
                            wb.save('SWD_data.xlsx')  # saving changes

                        else:
                            ws.cell(1, ws.max_column + 1).value = f'{values[0]}/{values[1]}'  # making new column
                            wb.save('SWD_data.xlsx')

                            for j in range(2, ws.max_row + 1):  # filling column -- none value for other personnel
                                if j == row + 2:
                                    ws.cell(j, ws.max_column).value = worked_minutes
                            wb.save('SWD_data.xlsx')  # saving changes

                    else:
                        sg.popup("Invalid Input", 'Input hours are wrong.')  # popup for invalid hours
            else:
                sg.popup("Invalid Input", "Invalid Date")  # popup fpr invalid dates

    window.close()


# password function, opening password window for manager access

def password():
    layout_password = [
        [sg.Text('password', size=(15, 1)), sg.InputText()],
        [sg.Submit()]
    ]
    window = sg.Window("Password", layout_password, size=(800, 300), modal=True)
    while True:
        event, values = window.read()  # input password into values list
        if event == sg.WIN_CLOSED:
            break
        if event == "Submit":
            if values[0] == '123':
                personnel_management_panel()  # entering management panel if the password is correct
                break
            else:
                sg.popup('Access Denied', 'Password in wrong')  # Access Denied popup if the password is wrong
    window.close()


# accessible for managers
# has salary calculator, register personnel, show all personnel information, delete and search options
def personnel_management_panel():
    layout_management = [
        [sg.Text('Welcome to Management Panel', font=('Arial Bold', 40), expand_x=True, justification='center')],
        [sg.Button("Salary Calculator", size=(85, 4), pad=(350, 10))],
        [sg.Button("Register Personnel", size=(85, 4), pad=(350, 10))],
        [sg.Button("Delete Personnel", size=(85, 4), pad=(350, 10))],
        [sg.Button("Show Personnel Information", size=(85, 4), pad=(350, 10))],
        [sg.Button("Search", size=(85, 4), pad=(350, 10))],
        [sg.Button("Back", size=(10, 4), pad=(600, 10))]
    ]
    window = sg.Window("Management Panel", layout_management, size=(1300, 700), modal=True)
    while True:
        event, values = window.read()
        if event == "Salary Calculator":
            salary_calculator()
        elif event == "Show Personnel Information":
            information_open_window()
        elif event == sg.WIN_CLOSED or event == 'Back':
            break
        elif event == "Register Personnel":
            register_open_window()
        elif event == "Search":
            search()
        elif event == "Delete Personnel":
            delete()
    window.close()


# delete function receives personnel ID and deletes that person with given OID and removes the row from Excel file and
# database. also deletes relevant data from PIDs and OIDs lists
def delete():
    layout_search = [
        [sg.Text('Enter Personnel ID:\n', size=(15, 1)), sg.InputText()],
        [sg.Submit('Delete')]
    ]
    window = sg.Window("Delete Personnel", layout_search, size=(600, 100), modal=True)
    while True:
        event, values = window.read()
        if event == sg.WIN_CLOSED:
            break
        if values[0] in OIDs:
            ws.delete_rows(idx=OIDs.index(values[0]) + 2)
            wb.save('SWD_data.xlsx')
            del PIDs[OIDs.index(values[0])]
            del OIDs[OIDs.index(values[0])]
            sg.popup("Personnel Deleted successfully!")
        else:
            sg.popup('Personnel ID doesnt exist')
    window.close()


# search function searches by personnel ID, ID, name, lastname or position
def search():
    df = pd.read_excel('SWD_data.xlsx')
    layout_search = [
        [sg.Text('Search by Personnel ID, Name, Lastname, Position or ID:\n', size=(45, 1)), sg.InputText()],
        [sg.Submit('Search')]
    ]
    window = sg.Window("Search Menu", layout_search, size=(600, 100), modal=True)
    while True:
        event, values = window.read()  # searched string is saved into values list
        if event == sg.WIN_CLOSED:
            break
        if event == 'Search':
            if values[0].isdigit():  # if the searched string is all digit, searches among OIDs and PIDs lists
                # and gets its index
                if values[0] in OIDs:
                    index = OIDs.index(values[0])
                    sg.popup_scrolled('Search Results:', f'Name: {df.loc[index][0]}  Lastname: {df.loc[index][1]}  '
                                                         f'ID: {ws.cell(row=index + 2, column=3).value}'
                                                         f'  Personnel ID: {ws.cell(row=index + 2, column=4).value}  '
                                                         f'Position:'
                                                         f' {df.loc[index][4]}')
                elif values[0] in PIDs:
                    index = PIDs.index(values[0])
                    sg.popup_scrolled('Search Results:', f'Name: {df.loc[index][0]}  Lastname: {df.loc[index][1]}  ID: '
                                                         f'{ws.cell(row=index + 2, column=3).value}'
                                                         f'  Personnel ID: {ws.cell(row=index + 2, column=4).value}  '
                                                         f'Position:'
                                                         f' {df.loc[index][4]}')
                else:
                    sg.popup_scrolled('Search Results:', 'No result was found :(')

            else:  # if searched string is not all digit, searches if the string contains every name, lastname
                # (not case-sensitive)
                index_list = []
                for index in range(df.shape[0]):
                    if df.loc[index]["Name"].lower() == str(values[0]).lower() or df.loc[index]["Lastname"].lower() \
                            == str(values[0]).lower() or df.loc[index]["Position"].lower() == str(values[0]).lower():
                        index_list.append(index)
                if len(index_list) == 0:
                    sg.popup_scrolled('Search Results:', 'No search results found :(')
                else:
                    newline = '\n'
                    sg.PopupScrolled("Personnel Data:",
                                     f"{newline.join(f' Name: {df.loc[j][0]}---Lastname: {df.loc[j][1]}---ID: {ws.cell(row=(j + 2), column=3).value}---Personnel ID: {ws.cell(row=(j + 2), column=4).value}---Position: {df.loc[j][4]}' for j in index_list)}")

    window.close()


# salary calculator function calculates salary based on working hours every month
# different salaries based on position and overtime work

def salary_calculator():
    layout_salary = [
        [sg.Text('Personnel ID', size=(15, 1)), sg.InputText()],
        [sg.Text('Month', size=(15, 1)), sg.InputText()],
        [sg.Submit()]
    ]
    window = sg.Window("Salary Calculator", layout_salary, size=(1000, 500), modal=True)
    while True:
        event, values = window.read()  # Personnel ID and number of month stores into values list
        if event == sg.WIN_CLOSED:
            break
        elif event == "Submit":
            # checking if values are valid
            if values[0] in OIDs:
                if values[1].isdigit() and 0 < int(values[1]) < 13:
                    day_index = []  # days worked this month gets stored in this list
                    for j in range(ws.max_column):
                        cell = ws.cell(row=1, column=j + 1).value
                        if cell[-2:].isdigit() and int(cell[-2:]) == int(values[1]) or cell[-1].isdigit() and int(
                                cell[-1]) == int(values[1]):
                            day_index.append(j + 1)
                    if values[0] in OIDs:
                        total_hours = 0
                        for j in range(len(day_index)):  # calculating total hours worked this month
                            if ws.cell(row=OIDs.index(values[0]) + 2, column=day_index[j]).value is not None:
                                total_hours += int(ws.cell(row=OIDs.index(values[0]) + 2, column=day_index[j]).value)
                        total_hours /= 60
                        total_hours = round(total_hours, 2)
                        position = ws.cell(row=OIDs.index(values[0]) + 2, column=5).value
                        match position:  # calculating salary based on position e.g. CEO doesn't receive overtime
                            # payment
                            case 'CEO':
                                payment = round(total_hours * 510)
                                sg.popup(
                                    '',
                                    f'{ws.cell(row=OIDs.index(values[0]) + 2, column=1).value} '
                                    f'{ws.cell(row=OIDs.index(values[0]) + 2, column=2).value}\nPosition: {position}\n'
                                    f'Total'
                                    f' working hours this month: {total_hours}\nThis Month Salary: ${payment}')
                            case 'CTO':
                                over_time = total_hours - 160
                                if over_time < 0:
                                    payment = round(total_hours * 140)
                                else:
                                    payment = round(160 * 140 + over_time * 150)
                                sg.popup(
                                    '',
                                    f'{ws.cell(row=OIDs.index(values[0]) + 2, column=1).value} '
                                    f'{ws.cell(row=OIDs.index(values[0]) + 2, column=2).value}\nPosition: {position}\n'
                                    f'Total'
                                    f' working hours this month: {total_hours}\nThis Month Salary: ${payment}')
                            case 'COO':
                                over_time = total_hours - 160
                                if over_time < 0:
                                    payment = round(total_hours * 170)
                                else:
                                    payment = round(160 * 170 + over_time * 180)
                                sg.popup(
                                    '',
                                    f'{ws.cell(row=OIDs.index(values[0]) + 2, column=1).value} '
                                    f'{ws.cell(row=OIDs.index(values[0]) + 2, column=2).value}\nPosition: {position}\n'
                                    f'Total'
                                    f' working hours this month: {total_hours}\nThis Month Salary: ${payment}')
                            case 'CMO':
                                over_time = total_hours - 160
                                if over_time < 0:
                                    payment = round(total_hours * 60)
                                else:
                                    payment = round(160 * 60 + over_time * 70)
                                sg.popup(
                                    '',
                                    f'{ws.cell(row=OIDs.index(values[0]) + 2, column=1).value} '
                                    f'{ws.cell(row=OIDs.index(values[0]) + 2, column=2).value}\nPosition: {position}\n'
                                    f'Total'
                                    f' working hours this month: {total_hours}\nThis Month Salary: ${payment}')
                            case 'Accountant':
                                over_time = total_hours - 160
                                if over_time < 0:
                                    payment = round(total_hours * 30)
                                else:
                                    payment = round(160 * 30 + over_time * 40)
                                sg.popup(
                                    '',
                                    f'{ws.cell(row=OIDs.index(values[0]) + 2, column=1).value} '
                                    f'{ws.cell(row=OIDs.index(values[0]) + 2, column=2).value}\nPosition: {position}\n'
                                    f'Total'
                                    f' working hours this month: {total_hours}\nThis Month Salary: ${payment}')
                else:
                    sg.popup("month input is invalid")
            else:
                sg.popup("No personnel with this personnel ID was found")

    window.close()


# register new personnel into database -- done by manager
# each register has 5 parameters
# Name, Lastname which should all be in English alphabet
# ID: 10-digit long. cant register a new person with ID that already exists
# Personnel ID: 5-digit long. cant register a new person with Personnel ID that already exists
# Position which should exist in positions list
def register_open_window():
    layout_register = [
        [sg.Text('Register Personnel', font=('Arial Bold', 40), expand_x=True, justification='left')],
        [sg.Text('Name', size=(15, 1)), sg.InputText()],
        [sg.Text('Lastname', size=(15, 1)), sg.InputText()],
        [sg.Text('ID', size=(15, 1)), sg.InputText()],
        [sg.Text('Personnel ID', size=(15, 1)), sg.InputText()],
        [sg.Text('Position', size=(15, 1)), sg.InputText()],
        [sg.Submit()]
    ]
    window = sg.Window("Office Management", layout_register, size=(1300, 700), modal=True)
    while True:
        event, values = window.read()
        if event == "Submit":
            if values[0].isalpha() and values[0] != "":
                if values[1].isalpha() and values[1] != "":
                    if values[2].isdigit() and values[2] != "" and len(values[2]) == 10:
                        if PIDs.count(values[2]) == 0:
                            if values[3].isdigit() and values[3] != "" and len(values[3]) == 5:
                                if OIDs.count(values[3]) == 0:
                                    if values[4] in list_of_positions:
                                        sg.popup("", "Personnel Data Submitted Successfully!")
                                        row = (values[0], values[1], str(values[2]), str(values[3]), values[4])
                                        ws.append(row)  # adding new row for newly registered person
                                        wb.save('SWD_data.xlsx')  # saving changes
                                        PIDs.append(values[2])  # adding new IDs to lists
                                        OIDs.append(values[3])
                                    else:
                                        sg.popup("Invalid Input", "Position is not valid")
                                else:
                                    sg.popup("Invalid Input", "This Personnel ID already exists")
                            else:
                                sg.popup("Invalid Input", "Invalid Personnel ID")
                        else:
                            sg.popup("Invalid Input", "This ID already exists")
                    else:
                        sg.popup("Invalid Input", "ID is invalid")
                else:
                    sg.popup("Invalid Input", "Lastname is invalid")
            else:
                sg.popup("Invalid Input", "Name is invalid")
        if event == "Exit" or event == sg.WIN_CLOSED:
            break
    window.close()


# this function shows all information about every personnel in the database
def information_open_window():
    df = pd.read_excel('SWD_data.xlsx')  # using dataframe to view Excel file and read it
    newline = '\n'
    sg.PopupScrolled("Personnel Data:",
                     f"{newline.join(f'{j + 1}. Name: {df.loc[j][0]}---Lastname: {df.loc[j][1]}---ID: {ws.cell(row=(j + 2), column=3).value}---Personnel ID: {ws.cell(row=(j + 2), column=4).value}---Position: {df.loc[j][4]}' for j in range(df.shape[0]))}")


# main function to load main menu, main page
def main():
    window = sg.Window("Main Menu", layout_welcome, size=(1300, 700))
    while True:
        event, values = window.read()
        if event == "Exit" or event == sg.WIN_CLOSED:
            break
        elif event == "Personnel Management Panel":
            password()
        elif event == "Personnel Panel":
            login()
    window.close()


if __name__ == "__main__":
    main()
